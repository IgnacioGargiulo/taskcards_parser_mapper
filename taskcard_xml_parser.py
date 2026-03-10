#!/usr/bin/env python3
"""
Minimal Task Card XML parser.

Outputs only 3 Excel sheets:
1) TaskCard_Metadata
2) TaskCard_Actionable
3) TaskCard_Tables

Arguments:
--xml     Required. Input XML path.
--output  Optional. Output XLSX path. Default: <xml_stem>_extract.xlsx
"""

import argparse
import re
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook


CODE_RE = re.compile(r"\b(\d{2}-\d{2}-\d{2}-\d{3}-\d{3})\b")
TAG_TYPES = {"A", "W", "C", "N", "f"}


def normalize_ws(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def safe_xml_text(xml_path: Path) -> str:
    # Source files may contain HTML entity &nbsp; which is undefined in XML parser.
    return xml_path.read_text(encoding="utf-8", errors="replace").replace("&nbsp;", " ")


def parse_xml(xml_path: Path) -> ET.Element:
    return ET.fromstring(safe_xml_text(xml_path))


def text_of(node: ET.Element | None) -> str:
    if node is None:
        return ""
    return normalize_ws("".join(node.itertext()))


def extract_metadata(task_card: ET.Element) -> dict[str, str]:
    scalar_fields = [
        "Task_Card_Name",
        "Status",
        "Editor_Used",
        "Revision",
        "Revised_By",
        "TC_Description",
        "Type",
        "Category",
        "Chapter",
        "Section",
        "Paragraph",
        "Area",
        "Etops_Flag",
        "Corrosion_Flag",
        "SSID_Flag",
        "NDT_Flag",
        "ET_Flag",
        "MT_Flag",
        "PT_Flag",
        "UT_Flag",
        "Skill_Inspector",
    ]
    out: dict[str, str] = {}
    for field in scalar_fields:
        out[field] = text_of(task_card.find(field))
    return out


def parse_table_rows(table_node: ET.Element) -> list[list[str]]:
    rows: list[list[str]] = []
    for tr in table_node.findall(".//tr"):
        cells = [normalize_ws("".join(td.itertext())) for td in tr.findall("./th") + tr.findall("./td")]
        if any(cells):
            rows.append(cells)
    return rows


def extract_tables(task_card: ET.Element, task_card_name: str) -> list[dict]:
    table_rows: list[dict] = []
    items = task_card.findall("./Task_Card_Items/Task_Card_Item")
    for item in items:
        item_number = text_of(item.find("./Item_Number"))
        task_text = item.find("./Task_Card_Text")
        if task_text is None:
            continue

        table_no = 0
        for child in list(task_text):
            if child.tag != "table":
                continue
            table_no += 1
            rows = parse_table_rows(child)
            for row_idx, cells in enumerate(rows, start=1):
                table_rows.append(
                    {
                        "Task_Card_Name": task_card_name,
                        "Item_Number": item_number,
                        "Table_No": table_no,
                        "Row_No": row_idx,
                        "Cells": cells,
                    }
                )
    return table_rows


def extract_code(text: str) -> str:
    m = CODE_RE.search(text or "")
    return m.group(1) if m else ""


def extract_leading_task_code(text: str) -> str:
    m = re.match(r"^\s*TASK\s+(\d{2}-\d{2}-\d{2}-\d{3}-\d{3})\b", text or "", flags=re.IGNORECASE)
    return m.group(1) if m else ""


def split_aircraft_and_code(text: str) -> tuple[str, str]:
    clean = normalize_ws(text)
    if not clean:
        return "", ""

    if "|" in clean:
        left, right = [normalize_ws(x) for x in clean.split("|", 1)]
        code = extract_code(right) or extract_code(left)
        if code:
            left_no_code = normalize_ws(CODE_RE.sub(" ", left))
            aircraft_note = left_no_code if left_no_code and "task " not in left_no_code.lower() else ""
            return aircraft_note, code
        return left, ""

    code = extract_code(clean)
    if code:
        aircraft_note = normalize_ws(CODE_RE.sub(" ", clean))
        if "task " in aircraft_note.lower():
            aircraft_note = ""
        return aircraft_note, code
    return clean, ""


def extract_structured_rows(task_card: ET.Element, task_card_name: str) -> list[dict]:
    rows: list[dict] = []
    items = task_card.findall("./Task_Card_Items/Task_Card_Item")

    for item in items:
        item_number = text_of(item.find("./Item_Number"))
        task_text = item.find("./Task_Card_Text")
        if task_text is None:
            continue

        section_no = ""
        section_title = ""
        step_no = ""
        substep_no = ""
        order_no = ""
        current_t = ""
        item_task_code = ""
        item_description = ""
        current_aircraft_note = ""
        current_subtask_code = ""
        seq = 0
        expecting_section_title = False

        for child in list(task_text):
            tag = child.tag
            if tag == "n":
                n_val = normalize_ws(child.text or "")
                if n_val:
                    if re.match(r"^[A-Z]\.$", n_val):
                        section_no = n_val
                        section_title = ""
                        step_no = ""
                        substep_no = ""
                        expecting_section_title = True
                    elif re.match(r"^\(\d+\)$", n_val):
                        step_no = n_val
                        substep_no = ""
                    elif re.match(r"^\([a-z]\)$", n_val):
                        substep_no = n_val
                    elif re.match(r"^\d+$", n_val):
                        order_no = n_val
                continue

            if tag == "t":
                current_t = normalize_ws(child.text or "")
                tail = normalize_ws(child.tail or "")
                if not tail:
                    continue

                seq += 1
                aircraft_note = current_aircraft_note
                subtask_code = current_subtask_code
                line_kind = "text"
                text = tail

                if current_t == "A":
                    line_kind = "aircraft_or_subtask"
                    aircraft, code = split_aircraft_and_code(tail)
                    if aircraft:
                        aircraft_note = aircraft
                        current_aircraft_note = aircraft
                    if code:
                        subtask_code = code
                        current_subtask_code = code
                    text = normalize_ws(CODE_RE.sub("", tail).replace("|", " "))
                elif current_t in TAG_TYPES:
                    line_kind = f"tag_{current_t}"

                if not current_t and expecting_section_title and not section_title:
                    section_title = tail
                    expecting_section_title = False
                    line_kind = "section_title"

                lead_task_code = extract_leading_task_code(tail)
                if lead_task_code:
                    item_task_code = lead_task_code
                    subtask_code = lead_task_code
                    current_subtask_code = lead_task_code

                rows.append(
                    {
                        "Task_Card_Name": task_card_name,
                        "Item_Number": item_number,
                        "Seq_No": seq,
                        "Section_No": section_no,
                        "Section_Title": section_title,
                        "Order_No": order_no,
                        "Step_No": step_no,
                        "Substep_No": substep_no,
                        "Tag_Type": current_t,
                        "Line_Kind": line_kind,
                        "Aircraft_Note": aircraft_note,
                        "Subtask_Code": subtask_code,
                        "Item_Task_Code": item_task_code,
                        "Item_Description": item_description,
                        "Text": text,
                    }
                )
                continue

            if tag == "p":
                text = normalize_ws("".join(child.itertext()))
                if not text:
                    continue

                seq += 1
                line_kind = "text"
                aircraft_note = current_aircraft_note
                subtask_code = current_subtask_code

                if expecting_section_title and not section_title:
                    section_title = text
                    expecting_section_title = False
                    line_kind = "section_title"
                elif current_t == "A":
                    line_kind = "aircraft_or_subtask"
                    aircraft, code = split_aircraft_and_code(text)
                    if aircraft:
                        aircraft_note = aircraft
                        current_aircraft_note = aircraft
                    if code:
                        subtask_code = code
                        current_subtask_code = code
                elif current_t in TAG_TYPES:
                    line_kind = f"tag_{current_t}"

                lead_task_code = extract_leading_task_code(text)
                if lead_task_code:
                    item_task_code = lead_task_code
                    subtask_code = lead_task_code
                    current_subtask_code = lead_task_code

                code_in_text = extract_code(text)
                if code_in_text and not subtask_code:
                    subtask_code = code_in_text
                    current_subtask_code = code_in_text

                if item_task_code and not item_description and not section_no:
                    item_description = text

                rows.append(
                    {
                        "Task_Card_Name": task_card_name,
                        "Item_Number": item_number,
                        "Seq_No": seq,
                        "Section_No": section_no,
                        "Section_Title": section_title,
                        "Order_No": order_no,
                        "Step_No": step_no,
                        "Substep_No": substep_no,
                        "Tag_Type": current_t,
                        "Line_Kind": line_kind,
                        "Aircraft_Note": aircraft_note,
                        "Subtask_Code": subtask_code,
                        "Item_Task_Code": item_task_code,
                        "Item_Description": item_description,
                        "Text": text,
                    }
                )
                continue

            if tag == "table":
                table_lines = []
                for cells in parse_table_rows(child):
                    table_lines.append("\t".join(normalize_ws(cell) for cell in cells))
                table_text = "\n".join(line for line in table_lines if line.strip()) or "[TABLE]"
                seq += 1
                rows.append(
                    {
                        "Task_Card_Name": task_card_name,
                        "Item_Number": item_number,
                        "Seq_No": seq,
                        "Section_No": section_no,
                        "Section_Title": section_title,
                        "Order_No": order_no,
                        "Step_No": step_no,
                        "Substep_No": substep_no,
                        "Tag_Type": current_t,
                        "Line_Kind": "table_text",
                        "Aircraft_Note": current_aircraft_note,
                        "Subtask_Code": current_subtask_code,
                        "Item_Task_Code": item_task_code,
                        "Item_Description": item_description,
                        "Text": table_text,
                    }
                )
                continue

    return rows


def build_actionable_rows(structured_rows: list[dict]) -> list[dict]:
    item_rows: list[dict] = []
    seen_items: set[tuple[str, str]] = set()
    for row in structured_rows:
        task = row.get("Task_Card_Name", "")
        item = row.get("Item_Number", "")
        item_desc = normalize_ws(str(row.get("Item_Description", "")))
        if not item_desc:
            continue
        key = (task, item)
        if key in seen_items:
            continue
        seen_items.add(key)
        item_rows.append(
            {
                "Task_Card_Name": task,
                "Item_Number": item,
                "Section_No": "",
                "Section_Title": "",
                "Step_No": "",
                "Substep_No": "",
                "Tag_Type": "",
                "Aircraft_Note": row.get("Aircraft_Note", ""),
                "Subtask_Code": row.get("Item_Task_Code", "") or row.get("Subtask_Code", ""),
                "Text": item_desc,
                "_Sort_Seq": int(row.get("Seq_No", 0) or 0),
            }
        )

    groups: dict[tuple, dict] = {}
    for row in structured_rows:
        line_kind = row.get("Line_Kind", "")
        if line_kind in {"section_title", "aircraft_or_subtask"}:
            continue
        if not row.get("Step_No") and not row.get("Substep_No"):
            continue

        key = (
            row.get("Task_Card_Name", ""),
            row.get("Item_Number", ""),
            row.get("Section_No", ""),
            row.get("Section_Title", ""),
            row.get("Step_No", ""),
            row.get("Substep_No", ""),
        )
        if key not in groups:
            groups[key] = {
                "Task_Card_Name": row.get("Task_Card_Name", ""),
                "Item_Number": row.get("Item_Number", ""),
                "Section_No": row.get("Section_No", ""),
                "Section_Title": row.get("Section_Title", ""),
                "Step_No": row.get("Step_No", ""),
                "Substep_No": row.get("Substep_No", ""),
                "Tag_Type_Set": set(),
                "Aircraft_Note": row.get("Aircraft_Note", ""),
                "Subtask_Code": row.get("Subtask_Code", ""),
                "Texts": [],
                "_Sort_Seq": int(row.get("Seq_No", 0) or 0),
            }

        g = groups[key]
        row_seq = int(row.get("Seq_No", 0) or 0)
        g["_Sort_Seq"] = min(g["_Sort_Seq"], row_seq)

        tag_type = normalize_ws(str(row.get("Tag_Type", "")))
        if tag_type:
            g["Tag_Type_Set"].add(tag_type)
        if not g["Aircraft_Note"]:
            g["Aircraft_Note"] = row.get("Aircraft_Note", "")
        if not g["Subtask_Code"]:
            g["Subtask_Code"] = row.get("Subtask_Code", "")

        text_raw = str(row.get("Text", ""))
        text_clean = text_raw if line_kind == "table_text" else normalize_ws(text_raw)
        if text_clean:
            g["Texts"].append(text_clean)

    merged_rows: list[dict] = []
    for g in groups.values():
        unique_texts: list[str] = []
        seen_texts: set[str] = set()
        for text in g["Texts"]:
            dedupe_key = normalize_ws(text)
            if dedupe_key and dedupe_key not in seen_texts:
                seen_texts.add(dedupe_key)
                unique_texts.append(text)
        merged_rows.append(
            {
                "Task_Card_Name": g["Task_Card_Name"],
                "Item_Number": g["Item_Number"],
                "Section_No": g["Section_No"],
                "Section_Title": g["Section_Title"],
                "Step_No": g["Step_No"],
                "Substep_No": g["Substep_No"],
                "Tag_Type": "|".join(sorted(g["Tag_Type_Set"])),
                "Aircraft_Note": g["Aircraft_Note"],
                "Subtask_Code": g["Subtask_Code"],
                "Text": "\n".join(unique_texts),
                "_Sort_Seq": g["_Sort_Seq"],
            }
        )

    all_rows = item_rows + merged_rows
    all_rows.sort(key=lambda r: (r["Task_Card_Name"], r["Item_Number"], r["_Sort_Seq"]))
    for row in all_rows:
        row.pop("_Sort_Seq", None)
    return all_rows


def write_excel(
    metadata: dict[str, str],
    actionable_rows: list[dict],
    tables: list[dict],
    output_path: Path,
) -> None:
    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = "TaskCard_Metadata"
    ws_meta.append(["Field", "Value"])
    for k, v in metadata.items():
        ws_meta.append([k, v])

    ws_action = wb.create_sheet("TaskCard_Actionable")
    action_headers = [
        "Task_Card_Name",
        "Item_Number",
        "Section_No",
        "Section_Title",
        "Step_No",
        "Substep_No",
        "Tag_Type",
        "Aircraft_Note",
        "Subtask_Code",
        "Text",
    ]
    ws_action.append(action_headers)
    for row in actionable_rows:
        ws_action.append([row.get(h, "") for h in action_headers])

    ws_tables = wb.create_sheet("TaskCard_Tables")
    max_cols = max((len(r["Cells"]) for r in tables), default=0)
    table_headers = ["Task_Card_Name", "Item_Number", "Table_No", "Row_No"] + [
        f"Col_{i}" for i in range(1, max_cols + 1)
    ]
    ws_tables.append(table_headers)
    for row in tables:
        cells = row["Cells"] + [""] * (max_cols - len(row["Cells"]))
        ws_tables.append(
            [row["Task_Card_Name"], row["Item_Number"], row["Table_No"], row["Row_No"]] + cells
        )

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse Task Card XML to a minimal Excel output.")
    parser.add_argument("--xml", required=True, type=Path, help="Input XML path")
    parser.add_argument("--output", type=Path, help="Output XLSX path")
    args = parser.parse_args()

    output_path = args.output or args.xml.with_name(f"{args.xml.stem}_extract.xlsx")

    root = parse_xml(args.xml)
    task_card = root.find(".//Task_Card_Element")
    if task_card is None:
        raise RuntimeError("Task_Card_Element not found in XML.")

    metadata = extract_metadata(task_card)
    task_card_name = metadata.get("Task_Card_Name", "")
    structured_rows = extract_structured_rows(task_card, task_card_name)
    actionable_rows = build_actionable_rows(structured_rows)
    table_rows = extract_tables(task_card, task_card_name)

    write_excel(metadata, actionable_rows, table_rows, output_path)
    print(
        f"Wrote {output_path} | metadata fields: {len(metadata)}, actionable rows: {len(actionable_rows)}, "
        f"table rows: {len(table_rows)}"
    )


if __name__ == "__main__":
    main()
