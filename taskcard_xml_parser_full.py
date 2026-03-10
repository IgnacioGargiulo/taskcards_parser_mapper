#!/usr/bin/env python3
"""
Full Task Card XML parser (pre-minimal variant).

Outputs Excel sheets:
- TaskCard_Metadata
- TaskCard_Text
- TaskCard_Structured
- TaskCard_Actionable
- TaskCard_Tables

Optional features:
- --pretty-xml
- --pdf (with --qa-report)
"""

import argparse
import json
import re
import subprocess
import xml.etree.ElementTree as ET
from pathlib import Path
from xml.dom import minidom

from openpyxl import Workbook


CODE_RE = re.compile(r"\b(\d{2}-\d{2}-\d{2}-\d{3}-\d{3})\b")
TAG_TYPES = {"A", "W", "C", "N", "f"}


def normalize_ws(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def safe_xml_text(xml_path: Path) -> str:
    return xml_path.read_text(encoding="utf-8", errors="replace").replace("&nbsp;", " ")


def parse_xml(xml_path: Path) -> ET.Element:
    return ET.fromstring(safe_xml_text(xml_path))


def write_pretty_xml(xml_path: Path, output_path: Path) -> None:
    pretty = minidom.parseString(safe_xml_text(xml_path)).toprettyxml(indent="  ", newl="\n")
    compact_pretty = "\n".join(line for line in pretty.splitlines() if line.strip()) + "\n"
    output_path.write_text(compact_pretty, encoding="utf-8")


def text_of(node: ET.Element | None) -> str:
    if node is None:
        return ""
    return normalize_ws("".join(node.itertext()))


def extract_metadata(task_card: ET.Element) -> dict[str, str]:
    scalar_fields = [
        "Task_Card_Name",
        "Status",
        "Editor_Used",
        "Revision",
        "Revised_By",
        "TC_Description",
        "Type",
        "Category",
        "Chapter",
        "Section",
        "Paragraph",
        "Area",
        "Etops_Flag",
        "Corrosion_Flag",
        "SSID_Flag",
        "NDT_Flag",
        "ET_Flag",
        "MT_Flag",
        "PT_Flag",
        "UT_Flag",
        "Skill_Inspector",
    ]
    out: dict[str, str] = {}
    for field in scalar_fields:
        out[field] = text_of(task_card.find(field))
    return out


def parse_table_rows(table_node: ET.Element) -> list[list[str]]:
    rows: list[list[str]] = []
    for tr in table_node.findall(".//tr"):
        cells = [normalize_ws("".join(td.itertext())) for td in tr.findall("./th") + tr.findall("./td")]
        if any(cells):
            rows.append(cells)
    return rows


def extract_code(text: str) -> str:
    m = CODE_RE.search(text or "")
    return m.group(1) if m else ""


def extract_leading_task_code(text: str) -> str:
    m = re.match(r"^\s*TASK\s+(\d{2}-\d{2}-\d{2}-\d{3}-\d{3})\b", text or "", flags=re.IGNORECASE)
    return m.group(1) if m else ""


def split_aircraft_and_code(text: str) -> tuple[str, str]:
    clean = normalize_ws(text)
    if not clean:
        return "", ""
    if "|" in clean:
        left, right = [normalize_ws(x) for x in clean.split("|", 1)]
        code = extract_code(right) or extract_code(left)
        if code:
            left_no_code = normalize_ws(CODE_RE.sub(" ", left))
            aircraft_note = left_no_code if left_no_code and "task " not in left_no_code.lower() else ""
            return aircraft_note, code
        return left, ""
    code = extract_code(clean)
    if code:
        aircraft_note = normalize_ws(CODE_RE.sub(" ", clean))
        if "task " in aircraft_note.lower():
            aircraft_note = ""
        return aircraft_note, code
    return clean, ""


def extract_structured_rows(task_card: ET.Element, task_card_name: str) -> list[dict]:
    rows: list[dict] = []
    items = task_card.findall("./Task_Card_Items/Task_Card_Item")

    for item in items:
        item_number = text_of(item.find("./Item_Number"))
        task_text = item.find("./Task_Card_Text")
        if task_text is None:
            continue

        section_no = ""
        section_title = ""
        step_no = ""
        substep_no = ""
        order_no = ""
        current_t = ""
        item_task_code = ""
        item_description = ""
        current_aircraft_note = ""
        current_subtask_code = ""
        seq = 0
        expecting_section_title = False

        for child in list(task_text):
            tag = child.tag
            if tag == "n":
                n_val = normalize_ws(child.text or "")
                if n_val:
                    if re.match(r"^[A-Z]\.$", n_val):
                        section_no = n_val
                        section_title = ""
                        step_no = ""
                        substep_no = ""
                        expecting_section_title = True
                    elif re.match(r"^\(\d+\)$", n_val):
                        step_no = n_val
                        substep_no = ""
                    elif re.match(r"^\([a-z]\)$", n_val):
                        substep_no = n_val
                    elif re.match(r"^\d+$", n_val):
                        order_no = n_val
                continue

            if tag == "t":
                current_t = normalize_ws(child.text or "")
                tail = normalize_ws(child.tail or "")
                if not tail:
                    continue

                seq += 1
                aircraft_note = current_aircraft_note
                subtask_code = current_subtask_code
                line_kind = "text"
                text = tail

                if current_t == "A":
                    line_kind = "aircraft_or_subtask"
                    aircraft, code = split_aircraft_and_code(tail)
                    if aircraft:
                        aircraft_note = aircraft
                        current_aircraft_note = aircraft
                    if code:
                        subtask_code = code
                        current_subtask_code = code
                    text = normalize_ws(CODE_RE.sub("", tail).replace("|", " "))
                elif current_t in TAG_TYPES:
                    line_kind = f"tag_{current_t}"

                if not current_t and expecting_section_title and not section_title:
                    section_title = tail
                    expecting_section_title = False
                    line_kind = "section_title"

                lead_task_code = extract_leading_task_code(tail)
                if lead_task_code:
                    item_task_code = lead_task_code
                    subtask_code = lead_task_code
                    current_subtask_code = lead_task_code

                rows.append(
                    {
                        "Task_Card_Name": task_card_name,
                        "Item_Number": item_number,
                        "Seq_No": seq,
                        "Section_No": section_no,
                        "Section_Title": section_title,
                        "Order_No": order_no,
                        "Step_No": step_no,
                        "Substep_No": substep_no,
                        "Tag_Type": current_t,
                        "Line_Kind": line_kind,
                        "Aircraft_Note": aircraft_note,
                        "Subtask_Code": subtask_code,
                        "Item_Task_Code": item_task_code,
                        "Item_Description": item_description,
                        "Text": text,
                        "Source": "tail_after_t",
                    }
                )
                continue

            if tag == "p":
                text = normalize_ws("".join(child.itertext()))
                if not text:
                    continue

                seq += 1
                line_kind = "text"
                aircraft_note = current_aircraft_note
                subtask_code = current_subtask_code

                if expecting_section_title and not section_title:
                    section_title = text
                    expecting_section_title = False
                    line_kind = "section_title"
                elif current_t == "A":
                    line_kind = "aircraft_or_subtask"
                    aircraft, code = split_aircraft_and_code(text)
                    if aircraft:
                        aircraft_note = aircraft
                        current_aircraft_note = aircraft
                    if code:
                        subtask_code = code
                        current_subtask_code = code
                elif current_t in TAG_TYPES:
                    line_kind = f"tag_{current_t}"

                lead_task_code = extract_leading_task_code(text)
                if lead_task_code:
                    item_task_code = lead_task_code
                    subtask_code = lead_task_code
                    current_subtask_code = lead_task_code

                code_in_text = extract_code(text)
                if code_in_text and not subtask_code:
                    subtask_code = code_in_text
                    current_subtask_code = code_in_text

                if item_task_code and not item_description and not section_no:
                    item_description = text

                rows.append(
                    {
                        "Task_Card_Name": task_card_name,
                        "Item_Number": item_number,
                        "Seq_No": seq,
                        "Section_No": section_no,
                        "Section_Title": section_title,
                        "Order_No": order_no,
                        "Step_No": step_no,
                        "Substep_No": substep_no,
                        "Tag_Type": current_t,
                        "Line_Kind": line_kind,
                        "Aircraft_Note": aircraft_note,
                        "Subtask_Code": subtask_code,
                        "Item_Task_Code": item_task_code,
                        "Item_Description": item_description,
                        "Text": text,
                        "Source": "p",
                    }
                )
                continue

            if tag == "table":
                table_lines = ["\t".join(normalize_ws(cell) for cell in cells) for cells in parse_table_rows(child)]
                table_text = "\n".join(line for line in table_lines if line.strip()) or "[TABLE]"
                seq += 1
                rows.append(
                    {
                        "Task_Card_Name": task_card_name,
                        "Item_Number": item_number,
                        "Seq_No": seq,
                        "Section_No": section_no,
                        "Section_Title": section_title,
                        "Order_No": order_no,
                        "Step_No": step_no,
                        "Substep_No": substep_no,
                        "Tag_Type": current_t,
                        "Line_Kind": "table_text",
                        "Aircraft_Note": current_aircraft_note,
                        "Subtask_Code": current_subtask_code,
                        "Item_Task_Code": item_task_code,
                        "Item_Description": item_description,
                        "Text": table_text,
                        "Source": "table",
                    }
                )
                continue

            extra_text = normalize_ws("".join(child.itertext()))
            if extra_text:
                seq += 1
                rows.append(
                    {
                        "Task_Card_Name": task_card_name,
                        "Item_Number": item_number,
                        "Seq_No": seq,
                        "Section_No": section_no,
                        "Section_Title": section_title,
                        "Order_No": order_no,
                        "Step_No": step_no,
                        "Substep_No": substep_no,
                        "Tag_Type": current_t,
                        "Line_Kind": tag,
                        "Aircraft_Note": current_aircraft_note,
                        "Subtask_Code": current_subtask_code,
                        "Item_Task_Code": item_task_code,
                        "Item_Description": item_description,
                        "Text": extra_text,
                        "Source": tag,
                    }
                )

    return rows


def extract_item_details(task_card: ET.Element, task_card_name: str) -> tuple[list[dict], list[dict]]:
    instruction_rows: list[dict] = []
    table_rows: list[dict] = []
    items = task_card.findall("./Task_Card_Items/Task_Card_Item")

    for item in items:
        item_number = text_of(item.find("./Item_Number"))
        task_text = item.find("./Task_Card_Text")
        if task_text is None:
            continue

        current_n = ""
        current_t = ""
        line_no = 0
        table_no = 0

        for child in list(task_text):
            tag = child.tag
            if tag == "n":
                current_n = normalize_ws(child.text or "")
                continue
            if tag == "t":
                current_t = normalize_ws(child.text or "")
                tail = normalize_ws(child.tail or "")
                if tail:
                    line_no += 1
                    instruction_rows.append(
                        {
                            "Task_Card_Name": task_card_name,
                            "Item_Number": item_number,
                            "Line_No": line_no,
                            "N_Tag": current_n,
                            "T_Tag": current_t,
                            "Text": tail,
                            "Source": "tail_after_t",
                        }
                    )
                continue
            if tag == "p":
                text = normalize_ws("".join(child.itertext()))
                if text:
                    line_no += 1
                    instruction_rows.append(
                        {
                            "Task_Card_Name": task_card_name,
                            "Item_Number": item_number,
                            "Line_No": line_no,
                            "N_Tag": current_n,
                            "T_Tag": current_t,
                            "Text": text,
                            "Source": "p",
                        }
                    )
                continue
            if tag == "table":
                table_no += 1
                table_data = parse_table_rows(child)
                if table_data:
                    for row_idx, cells in enumerate(table_data, start=1):
                        table_rows.append(
                            {
                                "Task_Card_Name": task_card_name,
                                "Item_Number": item_number,
                                "Table_No": table_no,
                                "Row_No": row_idx,
                                "Cells": cells,
                            }
                        )
                    line_no += 1
                    instruction_rows.append(
                        {
                            "Task_Card_Name": task_card_name,
                            "Item_Number": item_number,
                            "Line_No": line_no,
                            "N_Tag": current_n,
                            "T_Tag": current_t,
                            "Text": f"[TABLE {table_no}]",
                            "Source": "table_marker",
                        }
                    )
                continue
            extra_text = normalize_ws("".join(child.itertext()))
            if extra_text:
                line_no += 1
                instruction_rows.append(
                    {
                        "Task_Card_Name": task_card_name,
                        "Item_Number": item_number,
                        "Line_No": line_no,
                        "N_Tag": current_n,
                        "T_Tag": current_t,
                        "Text": extra_text,
                        "Source": tag,
                    }
                )

    return instruction_rows, table_rows


def normalize_for_compare(text: str) -> str:
    s = (text or "").lower()
    s = re.sub(r"#anchor#.*?#anchor#", " ", s)
    s = s.replace("\u00a0", " ")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def unique_preserve_order(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for v in values:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


def extract_pdf_text(pdf_path: Path) -> str:
    result = subprocess.run(
        ["pdftotext", "-layout", str(pdf_path), "-"],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return result.stdout


def build_qa_report(xml_path: Path, pdf_path: Path, metadata: dict[str, str], instructions: list[dict], tables: list[dict]) -> dict:
    xml_lines = [r["Text"] for r in instructions if r.get("Text") and not str(r.get("Text")).startswith("[TABLE")]
    xml_norm = unique_preserve_order([normalize_for_compare(x) for x in xml_lines if normalize_for_compare(x)])

    pdf_text = extract_pdf_text(pdf_path)
    pdf_norm_text = normalize_for_compare(pdf_text)
    pdf_lines = [ln.strip() for ln in pdf_text.splitlines() if ln.strip()]
    pdf_norm_lines = unique_preserve_order([n for n in (normalize_for_compare(ln) for ln in pdf_lines) if n and len(n) >= 12])

    xml_not_found = [x for x in xml_norm if x not in pdf_norm_text]
    xml_not_found_detailed: list[dict] = []
    for line in xml_not_found:
        toks = [t for t in line.split() if len(t) >= 3]
        hits = sum(1 for t in toks if t in pdf_norm_text)
        cov = (hits / len(toks)) if toks else 0.0
        if cov >= 0.9:
            cls = "formatting_wrap_artifact"
        elif cov >= 0.7:
            cls = "likely_present_split"
        else:
            cls = "likely_true_missing"
        xml_not_found_detailed.append({"line": line, "token_coverage": round(cov, 3), "classification": cls})

    xml_join = "\n".join(xml_norm)
    pdf_not_covered: list[str] = []
    for line in pdf_norm_lines:
        if line in xml_join:
            continue
        toks = [t for t in line.split() if len(t) >= 4]
        if len(toks) < 3:
            continue
        hits = sum(1 for t in toks if t in xml_join)
        if (hits / len(toks)) < 0.6:
            pdf_not_covered.append(line)

    return {
        "xml_file": str(xml_path),
        "pdf_file": str(pdf_path),
        "task_card_name": metadata.get("Task_Card_Name", ""),
        "counts": {
            "metadata_fields": len(metadata),
            "instruction_rows": len(instructions),
            "table_rows": len(tables),
            "xml_instruction_lines_total": len(xml_lines),
            "xml_instruction_lines_unique_norm": len(xml_norm),
            "pdf_lines_total_nonblank": len(pdf_lines),
            "pdf_lines_unique_norm_12plus": len(pdf_norm_lines),
            "xml_lines_not_found_in_pdf_count": len(xml_not_found),
            "pdf_lines_not_well_covered_by_xml_count": len(pdf_not_covered),
        },
        "classification_counts": {
            "formatting_wrap_artifact": sum(1 for r in xml_not_found_detailed if r["classification"] == "formatting_wrap_artifact"),
            "likely_present_split": sum(1 for r in xml_not_found_detailed if r["classification"] == "likely_present_split"),
            "likely_true_missing": sum(1 for r in xml_not_found_detailed if r["classification"] == "likely_true_missing"),
        },
        "xml_lines_not_found_in_pdf": xml_not_found_detailed,
        "pdf_lines_not_well_covered_by_xml_sample": pdf_not_covered[:120],
    }


def write_qa_report(report: dict, output_path: Path) -> None:
    output_path.write_text(json.dumps(report, indent=2), encoding="utf-8")


def build_actionable_rows(structured_rows: list[dict]) -> list[dict]:
    item_rows: list[dict] = []
    seen_items: set[tuple[str, str]] = set()
    for row in structured_rows:
        task = row.get("Task_Card_Name", "")
        item = row.get("Item_Number", "")
        item_desc = normalize_ws(str(row.get("Item_Description", "")))
        if not item_desc:
            continue
        key = (task, item)
        if key in seen_items:
            continue
        seen_items.add(key)
        item_rows.append(
            {
                "Task_Card_Name": task,
                "Item_Number": item,
                "Seq_No": row.get("Seq_No", ""),
                "Section_No": "",
                "Section_Title": "",
                "Step_No": "",
                "Substep_No": "",
                "Tag_Type": "",
                "Instruction_Type": "item_description",
                "Aircraft_Note": row.get("Aircraft_Note", ""),
                "Subtask_Code": row.get("Item_Task_Code", "") or row.get("Subtask_Code", ""),
                "Text": item_desc,
            }
        )

    groups: dict[tuple, dict] = {}
    for row in structured_rows:
        line_kind = row.get("Line_Kind", "")
        if line_kind in {"section_title", "aircraft_or_subtask"}:
            continue
        if not row.get("Step_No") and not row.get("Substep_No"):
            continue

        key = (
            row.get("Task_Card_Name", ""),
            row.get("Item_Number", ""),
            row.get("Section_No", ""),
            row.get("Section_Title", ""),
            row.get("Step_No", ""),
            row.get("Substep_No", ""),
        )
        if key not in groups:
            groups[key] = {
                "Task_Card_Name": row.get("Task_Card_Name", ""),
                "Item_Number": row.get("Item_Number", ""),
                "Seq_No": row.get("Seq_No", 0),
                "Section_No": row.get("Section_No", ""),
                "Section_Title": row.get("Section_Title", ""),
                "Step_No": row.get("Step_No", ""),
                "Substep_No": row.get("Substep_No", ""),
                "Tag_Type_Set": set(),
                "Instruction_Type_Set": set(),
                "Aircraft_Note": row.get("Aircraft_Note", ""),
                "Subtask_Code": row.get("Subtask_Code", ""),
                "Texts": [],
            }

        g = groups[key]
        seq_no = row.get("Seq_No", 0)
        if isinstance(seq_no, int) and isinstance(g["Seq_No"], int):
            g["Seq_No"] = min(g["Seq_No"], seq_no)

        tag_type = normalize_ws(str(row.get("Tag_Type", "")))
        if tag_type:
            g["Tag_Type_Set"].add(tag_type)
        if line_kind:
            g["Instruction_Type_Set"].add(line_kind)
        if not g["Aircraft_Note"]:
            g["Aircraft_Note"] = row.get("Aircraft_Note", "")
        if not g["Subtask_Code"]:
            g["Subtask_Code"] = row.get("Subtask_Code", "")

        text_raw = str(row.get("Text", ""))
        text_clean = text_raw if line_kind == "table_text" else normalize_ws(text_raw)
        if text_clean:
            g["Texts"].append(text_clean)

    merged_rows: list[dict] = []
    for g in groups.values():
        uniq: list[str] = []
        seen: set[str] = set()
        for text in g["Texts"]:
            k = normalize_ws(text)
            if k and k not in seen:
                seen.add(k)
                uniq.append(text)
        merged_rows.append(
            {
                "Task_Card_Name": g["Task_Card_Name"],
                "Item_Number": g["Item_Number"],
                "Seq_No": g["Seq_No"],
                "Section_No": g["Section_No"],
                "Section_Title": g["Section_Title"],
                "Step_No": g["Step_No"],
                "Substep_No": g["Substep_No"],
                "Tag_Type": "|".join(sorted(g["Tag_Type_Set"])),
                "Instruction_Type": "merged",
                "Aircraft_Note": g["Aircraft_Note"],
                "Subtask_Code": g["Subtask_Code"],
                "Text": "\n".join(uniq),
            }
        )

    return item_rows + sorted(merged_rows, key=lambda r: (r["Task_Card_Name"], r["Item_Number"], r["Seq_No"]))


def write_excel(metadata: dict[str, str], instructions: list[dict], structured_rows: list[dict], tables: list[dict], output_path: Path) -> None:
    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = "TaskCard_Metadata"
    ws_meta.append(["Field", "Value"])
    for k, v in metadata.items():
        ws_meta.append([k, v])

    ws_steps = wb.create_sheet("TaskCard_Text")
    step_headers = ["Task_Card_Name", "Item_Number", "Line_No", "N_Tag", "T_Tag", "Text", "Source"]
    ws_steps.append(step_headers)
    for row in instructions:
        ws_steps.append([row.get(h, "") for h in step_headers])

    ws_struct = wb.create_sheet("TaskCard_Structured")
    struct_headers = [
        "Task_Card_Name", "Item_Number", "Seq_No", "Section_No", "Section_Title", "Order_No",
        "Step_No", "Substep_No", "Tag_Type", "Line_Kind", "Aircraft_Note", "Subtask_Code",
        "Item_Task_Code", "Item_Description", "Text", "Source",
    ]
    ws_struct.append(struct_headers)
    for row in structured_rows:
        ws_struct.append([row.get(h, "") for h in struct_headers])

    ws_action = wb.create_sheet("TaskCard_Actionable")
    action_headers = [
        "Task_Card_Name", "Item_Number", "Seq_No", "Section_No", "Section_Title", "Step_No",
        "Substep_No", "Tag_Type", "Instruction_Type", "Aircraft_Note", "Subtask_Code", "Text",
    ]
    ws_action.append(action_headers)
    for row in build_actionable_rows(structured_rows):
        ws_action.append([row.get(h, "") for h in action_headers])

    ws_tables = wb.create_sheet("TaskCard_Tables")
    max_cols = max((len(r["Cells"]) for r in tables), default=0)
    table_headers = ["Task_Card_Name", "Item_Number", "Table_No", "Row_No"] + [f"Col_{i}" for i in range(1, max_cols + 1)]
    ws_tables.append(table_headers)
    for row in tables:
        cells = row["Cells"] + [""] * (max_cols - len(row["Cells"]))
        ws_tables.append([row["Task_Card_Name"], row["Item_Number"], row["Table_No"], row["Row_No"]] + cells)

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Full Task Card XML parser with optional QA/pretty features.")
    parser.add_argument("--xml", required=True, type=Path, help="Input XML path")
    parser.add_argument("--output", type=Path, help="Output XLSX path")
    parser.add_argument("--pretty-xml", nargs="?", const="__AUTO__", default=None, type=str, help="Write pretty XML copy")
    parser.add_argument("--pdf", type=Path, help="Source PDF path for optional QA report")
    parser.add_argument("--qa-report", type=Path, help="Output JSON path for QA report")
    args = parser.parse_args()

    output_path = args.output or args.xml.with_name(f"{args.xml.stem}_extract.xlsx")

    root = parse_xml(args.xml)
    task_card = root.find(".//Task_Card_Element")
    if task_card is None:
        raise RuntimeError("Task_Card_Element not found in XML.")

    metadata = extract_metadata(task_card)
    task_card_name = metadata.get("Task_Card_Name", "")
    instructions, tables = extract_item_details(task_card, task_card_name)
    structured_rows = extract_structured_rows(task_card, task_card_name)

    write_excel(metadata, instructions, structured_rows, tables, output_path)
    msg = (
        f"Wrote {output_path} | metadata fields: {len(metadata)}, text lines: {len(instructions)}, "
        f"structured rows: {len(structured_rows)}, table rows: {len(tables)}"
    )

    if args.pretty_xml is not None:
        pretty_path = args.xml.with_name(f"{args.xml.stem}.pretty.xml") if args.pretty_xml == "__AUTO__" else Path(args.pretty_xml)
        write_pretty_xml(args.xml, pretty_path)
        msg += f" | Pretty XML: {pretty_path}"

    if args.pdf:
        qa_path = args.qa_report or args.xml.with_name(f"{args.xml.stem}_qa_report.json")
        report = build_qa_report(args.xml, args.pdf, metadata, instructions, tables)
        write_qa_report(report, qa_path)
        msg += f" | QA report: {qa_path}"

    print(msg)


if __name__ == "__main__":
    main()
